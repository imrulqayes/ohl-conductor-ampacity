"""
Excel Report Generator for Conductor Ampacity Calculations.

Creates a formatted Excel workbook with:
  - Summary sheet  : All inputs and key results at a glance
  - Transient sheet: Time-series temperature profile (if transient was calculated)

Author       : Imrul Qayes
Email        : imrul27@gmail.com
Date Created : 03/2025
Last Modified: 04/2026
Version      : 1.0.0
AI Assistant : Developed with the assistance of Claude (Anthropic) — https://claude.ai
"""

# Licensed under the MIT License. See LICENSE file in the project root for details.

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter


# ── Color palette ────────────────────────────────────────────────────────────
_C_TITLE_BG   = "1F4E79"   # dark blue
_C_SECTION_BG = "2E75B6"   # medium blue
_C_LABEL_BG   = "D6E4F0"   # light blue
_C_RESULT_BG  = "E2EFDA"   # light green
_C_WARNING_BG = "FCE4D6"   # light orange
_C_WHITE      = "FFFFFF"
_C_BLACK      = "000000"
_C_DARK_TEXT  = "1F3864"


def _font(bold=False, size=11, color=_C_BLACK, italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic)


def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _thin_border():
    side = Side(style="thin", color="AAAAAA")
    return Border(left=side, right=side, top=side, bottom=side)


def _set(cell, value=None, bold=False, size=11, color=_C_BLACK, fill_color=None,
         align_h="left", wrap=False, italic=False, border=False):
    if value is not None:
        cell.value = value
    cell.font = _font(bold=bold, size=size, color=color, italic=italic)
    cell.alignment = _align(h=align_h, wrap=wrap)
    if fill_color:
        cell.fill = _fill(fill_color)
    if border:
        cell.border = _thin_border()


def _merge_title(ws, row, col_start, col_end, text, size=11, fill_color=None, bold=False, color=_C_BLACK, italic=False):
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end
    )
    cell = ws.cell(row=row, column=col_start)
    _set(cell, value=text, bold=bold, size=size, color=color,
         fill_color=fill_color, align_h="center", italic=italic)


# ── Public API ───────────────────────────────────────────────────────────────

def create_report(output_folder: str | Path,
                  conductor: dict,
                  env: dict,
                  solar: dict,
                  optical: dict,
                  calc_settings: dict,
                  results: dict,
                  standard: str = "IEEE 738") -> Path:
    """
    Create a formatted Excel report and save it to output_folder.

    Parameters:
        output_folder  : Directory where the file will be saved
        conductor      : Conductor properties dict
        env            : Environmental parameters dict
        solar          : Solar parameters dict
        optical        : Emissivity / absorptivity dict
        calc_settings  : Calculation settings dict
        results        : Dict with keys 'thermal_rating', 'steady_temperature', 'transient'
        standard       : Calculation standard label for the report header

    Returns:
        Path to the saved Excel file
    """
    output_folder = Path(output_folder)
    output_folder.mkdir(parents=True, exist_ok=True)

    cname  = conductor.get("name", "conductor").replace(" ", "_")
    ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname  = output_folder / f"{cname}_{ts}.xlsx"

    wb = openpyxl.Workbook()

    _build_summary_sheet(wb.active, conductor, env, solar, optical, calc_settings, results, standard)

    if results.get("transient") is not None:
        ws_tran = wb.create_sheet("Transient Profile")
        _build_transient_sheet(ws_tran, results["transient"], calc_settings)

    wb.save(fname)
    return fname


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_summary_sheet(ws, conductor, env, solar, optical, calc_settings, results, standard):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    row = 1

    # ── Main title ─────────────────────────────────────────────────────────
    _merge_title(ws, row, 1, 4,
                 f"Conductor Ampacity Calculation  —  {standard}",
                 size=14, fill_color=_C_TITLE_BG, bold=True, color=_C_WHITE)
    ws.row_dimensions[row].height = 28
    row += 1

    calc_time = datetime.now().strftime("%d %b %Y  %H:%M:%S")
    _merge_title(ws, row, 1, 4,
                 f"Generated: {calc_time}",
                 size=9, fill_color=_C_TITLE_BG, color="BBCCDD", italic=True)
    ws.row_dimensions[row].height = 16
    row += 2

    # ── Conductor Information ──────────────────────────────────────────────
    row = _section_header(ws, row, "Conductor Information")
    row = _data_row(ws, row, "Conductor Name",          conductor.get("name", "—"))
    row = _data_row(ws, row, "Outside Diameter",        f"{conductor.get('diameter_mm', '—')} mm")
    mt = conductor.get("max_temp")
    row = _data_row(ws, row, "Max Allowable Temperature",
                    f"{mt:.0f} °C" if mt is not None else "—  (using max_conductor_temp_C from config.ini)")
    row = _data_row(ws, row, "AC Resistance (low ref.)", f"{conductor.get('r_low', 0)*1000:.4f} mOhm/m  "
                                                          f"@ {conductor.get('temp_low', 25):.0f} °C")
    row = _data_row(ws, row, "AC Resistance (high ref.)", f"{conductor.get('r_high', 0)*1000:.4f} mOhm/m  "
                                                           f"@ {conductor.get('temp_high', 75):.0f} °C")
    hc = conductor.get("heat_capacity")
    row = _data_row(ws, row, "Heat Capacity",           f"{hc} J/(m.°C)" if hc else "—  (not required for steady-state)")
    row += 1

    # ── Environmental Conditions ───────────────────────────────────────────
    row = _section_header(ws, row, "Environmental Conditions")
    row = _data_row(ws, row, "Elevation",               f"{env.get('elevation', 0)} m")
    row = _data_row(ws, row, "Ambient Temperature",     f"{env.get('ambient_temp', '—')} °C")
    row = _data_row(ws, row, "Wind Speed",              f"{env.get('wind_speed', '—')} m/s")
    row = _data_row(ws, row, "Wind-to-Conductor Angle", f"{env.get('wind_angle', '—')}°")
    row = _data_row(ws, row, "Atmosphere Type",         env.get("atmosphere", "clear").capitalize())
    row += 1

    # ── Solar Parameters ───────────────────────────────────────────────────
    row = _section_header(ws, row, "Solar Parameters")
    row = _data_row(ws, row, "Latitude",                f"{solar.get('latitude', '—')}°")
    row = _data_row(ws, row, "Date",                    solar.get("date", "—"))
    row = _data_row(ws, row, "Hour of Day",             f"{solar.get('hour', '—')}:00  (24-h format)")
    row = _data_row(ws, row, "Line Azimuth",            f"{solar.get('line_azimuth', '—')}°  (clockwise from North)")
    q_s = results.get("solar_heat_gain")
    row = _data_row(ws, row, "Solar Heat Gain (Qs)",
                    f"{q_s:.2f} W/m²" if q_s is not None else "—")
    row += 1

    # ── Conductor Surface Properties ───────────────────────────────────────
    row = _section_header(ws, row, "Conductor Surface Properties")
    row = _data_row(ws, row, "Emissivity",              optical.get("emissivity", "—"))
    row = _data_row(ws, row, "Solar Absorptivity",      optical.get("absorptivity", "—"))
    row += 1

    # ── Calculation Results ────────────────────────────────────────────────
    row = _section_header(ws, row, "Calculation Results")

    if results.get("thermal_rating") is not None:
        max_t = calc_settings.get("max_conductor_temp_C", "—")
        row = _result_row(ws, row,
                          "Steady-State Thermal Rating",
                          f"{results['thermal_rating']} A",
                          f"at max. conductor temp. {max_t} °C")

    if results.get("steady_temperature") is not None:
        cur = calc_settings.get("steady_current_A", "—")
        row = _result_row(ws, row,
                          "Steady-State Conductor Temperature",
                          f"{results['steady_temperature']} °C",
                          f"at {cur} A")

    if results.get("transient"):
        curves   = results["transient"]
        ini_amps = calc_settings.get("transient_pre_current_A", "—")
        ini_temp = curves[0]["ini_temp"]

        row = _section_header(ws, row, "Transient Analysis Results")
        # Sub-header: pre-fault info
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        _set(ws.cell(row=row, column=1),
             value=f"  Pre-fault current: {ini_amps} A    |    "
                   f"Initial conductor temperature: {ini_temp} °C",
             size=9, italic=True, fill_color=_C_LABEL_BG, color="444444")
        ws.row_dimensions[row].height = 14
        row += 1

        # Table column headers  (4 columns: level | current | temp@10min | peak@time)
        for col_off, label in enumerate(["Fault Level",
                                         "Current (A)",
                                         "Temp @ 10 min (°C)",
                                         "Peak Temp (°C)  @  Time (min)"]):
            _set(ws.cell(row=row, column=col_off + 1),
                 value=label, bold=True,
                 fill_color=_C_SECTION_BG, color=_C_WHITE,
                 align_h="center", border=True, wrap=True)
        ws.row_dimensions[row].height = 28
        row += 1

        # One row per fault level
        for i, curve in enumerate(curves):
            fill = _C_LABEL_BG if i % 2 == 0 else _C_WHITE
            t10  = _temp_at_time(curve, target_min=10)
            _set(ws.cell(row=row, column=1),
                 value=f"  {curve['pct']:.0f}%", bold=True,
                 fill_color=fill, border=True)
            _set(ws.cell(row=row, column=2),
                 value=f"{curve['tran_amps']:.0f}",
                 fill_color=fill, align_h="center", border=True)
            _set(ws.cell(row=row, column=3),
                 value=t10 if t10 is not None else "—",
                 fill_color=fill, align_h="center", border=True)
            _set(ws.cell(row=row, column=4),
                 value=f"{curve['peak_temp']} °C  @  {curve['peak_time_min']:.1f} min",
                 bold=True, fill_color=_C_RESULT_BG, color=_C_DARK_TEXT,
                 align_h="center", border=True)
            ws.row_dimensions[row].height = 16
            row += 1

    row += 1
    _merge_title(ws, row, 1, 4,
                 "See 'Transient Profile' sheet for time-series data." if results.get("transient") else "",
                 size=9, italic=True, fill_color=None, color="888888")


def _build_transient_sheet(ws, curves, calc_settings):
    """
    Build the Transient Profile sheet with one column per fault level.

    Layout:
        Col A       : Time (min)
        Col B onward: Temperature (°C) for each fault level (110%, 120%, ...)
    """
    ws.sheet_view.showGridLines = False

    ini_amps   = calc_settings.get("transient_pre_current_A", "—")
    rating_A   = calc_settings.get("transient_rating_A", "—")
    n_curves   = len(curves)
    n_cols     = n_curves + 1   # time + one per curve
    last_col_letter = get_column_letter(n_cols)

    # ── Column widths ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 16
    for c in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # ── Row 1: Title ───────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col_letter}1")
    _set(ws["A1"], value="Transient Temperature Profile",
         bold=True, size=13, fill_color=_C_TITLE_BG, color=_C_WHITE, align_h="center")
    ws.row_dimensions[1].height = 26

    # ── Row 2: Sub-title ───────────────────────────────────────────────────
    ws.merge_cells(f"A2:{last_col_letter}2")
    _set(ws["A2"],
         value=(f"Pre-fault current: {ini_amps} A    |    "
                f"Thermal rating: {rating_A} A    |    "
                f"Fault levels: % of rated current"),
         size=9, fill_color=_C_TITLE_BG, color="BBCCDD", align_h="center", italic=True)
    ws.row_dimensions[2].height = 15

    # ── Row 3: Column headers ──────────────────────────────────────────────
    _set(ws.cell(row=3, column=1), value="Time (min)",
         bold=True, fill_color=_C_SECTION_BG, color=_C_WHITE,
         align_h="center", border=True)
    for j, curve in enumerate(curves, start=2):
        _set(ws.cell(row=3, column=j),
             value=f"{curve['pct']:.0f}%  ({curve['tran_amps']:.0f} A)",
             bold=True, fill_color=_C_SECTION_BG, color=_C_WHITE,
             align_h="center", border=True, wrap=True)
    ws.row_dimensions[3].height = 22

    # All curves share the same time axis — use the first one
    time_min = [t / 60.0 for t in curves[0]["time_s"]]
    n_data   = len(time_min)

    # ── Data rows ──────────────────────────────────────────────────────────
    for i, t_min in enumerate(time_min):
        data_row = i + 4
        row_fill = _C_RESULT_BG if i % 2 == 0 else _C_WHITE

        _set(ws.cell(row=data_row, column=1),
             value=round(t_min, 4), align_h="center",
             fill_color=row_fill, border=True)

        for j, curve in enumerate(curves, start=2):
            temp     = curve["temp_c"][i]
            is_peak  = (temp == curve["peak_temp"])
            _set(ws.cell(row=data_row, column=j),
                 value=temp, align_h="center",
                 fill_color=_C_WARNING_BG if is_peak else row_fill,
                 border=True, bold=is_peak)

    # ── Scatter chart (all curves overlaid) ────────────────────────────────
    last_data_row = 3 + n_data   # row 3 = header, row 4 = first data point

    chart = ScatterChart()
    chart.style          = 2       # clean style — plain axes, no fill, values visible
    chart.roundedCorners = False   # sharp corners for compact look

    # ── Axis labels (no chart title — saves vertical space) ────────────────
    chart.x_axis.title = "Time (min)"
    chart.y_axis.title = "Temperature (°C)"

    # ── Axis number formats ────────────────────────────────────────────────
    chart.x_axis.numFmt = "0"
    chart.y_axis.numFmt = "0"

    # ── Axis tick label positions ──────────────────────────────────────────
    chart.x_axis.tickLblPos = "low"     # labels at bottom of chart
    chart.y_axis.tickLblPos = "nextTo"  # labels immediately beside the axis

    # ── Axes must be visible ───────────────────────────────────────────────
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    # ── X axis: fixed 0 – 60 min range, tick every 10 min ─────────────────
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 60
    chart.x_axis.majorUnit   = 10

    # ── Y axis: tight range fitted to the actual temperature data ──────────
    all_temps = [t for c in curves for t in c["temp_c"]]
    y_lo  = min(all_temps)
    y_hi  = max(all_temps)
    pad   = (y_hi - y_lo) * 0.06

    # Snap min/max to the nearest multiple of 10 with a small margin
    y_min = max(0, int((y_lo - pad) / 10) * 10)
    y_max = (int((y_hi + pad) / 10) + 1) * 10

    y_range = y_max - y_min
    if   y_range <= 50:  y_unit = 5
    elif y_range <= 150: y_unit = 10
    elif y_range <= 300: y_unit = 20
    else:                y_unit = 50

    chart.y_axis.scaling.min = y_min
    chart.y_axis.scaling.max = y_max
    chart.y_axis.majorUnit   = y_unit

    # ── Major grid lines on both axes ─────────────────────────────────────
    chart.x_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines = ChartLines()

    # ── Series: smooth coloured lines, no point markers ────────────────────
    _CHART_COLORS = ["2563EB", "16A34A", "D97706", "DC2626", "7C3AED"]
    xvalues = Reference(ws, min_col=1, min_row=4, max_row=last_data_row)

    for j, curve in enumerate(curves):
        yvalues = Reference(ws, min_col=j + 2, min_row=4, max_row=last_data_row)
        ser     = Series(yvalues, xvalues,
                         title=f"{curve['pct']:.0f}%  ({curve['tran_amps']:.0f} A)")
        ser.smooth = True
        ser.graphicalProperties.line.solidFill = _CHART_COLORS[j % len(_CHART_COLORS)]
        ser.graphicalProperties.line.width      = 25000   # ~1.9 pt
        ser.marker.symbol = "none"
        chart.series.append(ser)

    # ── Legend at right to maximise vertical plot height ───────────────────
    chart.legend.position = "r"

    # ── Chart size: 26 cm wide × 16 cm tall ───────────────────────────────
    chart.width  = 26
    chart.height = 16

    chart_anchor = f"A{last_data_row + 3}"
    ws.add_chart(chart, chart_anchor)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _temp_at_time(curve: dict, target_min: float):
    """Return the conductor temperature at *target_min* minutes (nearest time step)."""
    target_s = target_min * 60.0
    for i, t in enumerate(curve["time_s"]):
        if t >= target_s:
            return curve["temp_c"][i]
    return curve["temp_c"][-1]   # target beyond simulation range


# ── Row helpers ───────────────────────────────────────────────────────────────

def _section_header(ws, row, title):
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=4
    )
    cell = ws.cell(row=row, column=1)
    _set(cell, value=f"  {title}", bold=True, size=11,
         fill_color=_C_SECTION_BG, color=_C_WHITE)
    ws.row_dimensions[row].height = 18
    return row + 1


def _data_row(ws, row, label, value):
    c_label = ws.cell(row=row, column=1)
    c_value = ws.cell(row=row, column=2)

    _set(c_label, value=f"  {label}", fill_color=_C_LABEL_BG, border=True)
    _set(c_value, value=str(value), align_h="left", border=True)

    # merge value across remaining columns for readability
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 16
    return row + 1


def _result_row(ws, row, label, value, note=""):
    c_label = ws.cell(row=row, column=1)
    c_value = ws.cell(row=row, column=2)
    c_note  = ws.cell(row=row, column=3)

    _set(c_label, value=f"  {label}", fill_color=_C_LABEL_BG, bold=True, border=True)
    _set(c_value, value=str(value), bold=True, fill_color=_C_RESULT_BG,
         color=_C_DARK_TEXT, align_h="center", border=True, size=12)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    _set(c_note, value=str(note), size=9, color="555555", italic=True, border=True)

    ws.row_dimensions[row].height = 18
    return row + 1
